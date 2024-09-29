import openpyxl
import psycopg2
from config_db import *

def make_connection():
    return psycopg2.connect(
        host=host,
        user=user,
        password=password,
        dbname=dbname,
        port=port
    )

def get_keycloak_roles():
    c = make_connection()  # Создание подключения
    cursor = c.cursor()
    # Выполнение запроса
    cursor.execute("SELECT name, id FROM keycloak_roles;")
    keycloak_roles = cursor.fetchall()  # Получение всех результатов
    c.close()  # Закрываем соединение
    return keycloak_roles

def find_role_excel(keycloak_role):
    excel_file = "mapping.xlsx"  # Путь к Excel-файлу
    workbook = openpyxl.load_workbook(excel_file) # Загружаем Excel-файл
    sheet = workbook.active  # Выбираем активный лист
    keycloak_role = keycloak_role # Значение, которое ищем в столбце B

    keycloak_roles_excel = []  # Список для хранения всех значений из столбца G

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):  # Проход по всем строкам
        cell_b = row[1]  # Столбец B
        cell_g = row[6]  # Столбец G

        if cell_b.value == keycloak_role:  # Проверяем ячейки в столбце B
            keycloak_roles_excel.append(cell_g.value)
    return keycloak_roles_excel

def get_id_role_db():
    c = make_connection()  # Создание подключения
    cursor = c.cursor()
    # Выполнение запроса
    cursor.execute("SELECT name, id FROM roles WHERE status = 1;")
    roles_db = cursor.fetchall()  # Получение всех результатов
    c.close()  # Закрываем соединение
    return roles_db

roles_db = get_id_role_db()

keycloak_roles = get_keycloak_roles()

for keycloak_role in keycloak_roles:
    id_role_db = []
    keycloak_roles_excel = find_role_excel(keycloak_role[0])
    for keycloak_role_excel in keycloak_roles_excel:
        for role_db in roles_db:
            if keycloak_role_excel == role_db[0]:
                id_role_db.append(f"({keycloak_role[1]}, {role_db[1]})")
    if id_role_db:
        id_role_db_str = ", ".join(id_role_db)
        print(f"INSERT INTO keycloak_role_roles (keycloak_role_id, role_id) VALUES {id_role_db_str};")